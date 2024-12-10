import openpyxl
from openpyxl import Workbook
import os

class NhanVien:
    def __init__(self, ma_nv, ten, tuoi):
        self.ma_nv = ma_nv
        self.ten = ten
        self.tuoi = tuoi

    def __repr__(self):
        return f"NhanVien({self.ma_nv}, {self.ten}, {self.tuoi})"

class QuanLyNhanVien:
    def __init__(self):
        self.nhan_vien_list = []

    def them_nhan_vien(self, nhan_vien):
        self.nhan_vien_list.append(nhan_vien)

    def cap_nhat_nhan_vien(self, ma_nv, ten=None, tuoi=None):
        for nv in self.nhan_vien_list:
            if nv.ma_nv == ma_nv:
                if ten:
                    nv.ten = ten
                if tuoi:
                    nv.tuoi = tuoi
                return
        print("Không tìm thấy nhân viên")

    def xoa_nhan_vien(self, ma_nv):
        for nv in self.nhan_vien_list:
            if nv.ma_nv == ma_nv:
                self.nhan_vien_list.remove(nv)
                return
        print("Không tìm thấy nhân viên")

    def sap_xep_nhan_vien(self):
        self.nhan_vien_list.sort(key=lambda nv: nv.tuoi)

    def luu_vao_file(self, ten_file):
        if not ten_file.endswith(".xlsx"):
            ten_file += ".xlsx"

        current_directory = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_directory, ten_file)
        
        wb = Workbook()
        ws = wb.active
        ws.append(["Mã", "Tên", "Tuổi"])  # Đặt tiêu đề cột

        for nv in self.nhan_vien_list:
            ws.append([nv.ma_nv, nv.ten, nv.tuoi])

        wb.save(file_path)
        print(f"Dữ liệu đã được lưu vào: {file_path}")

    def doc_tu_file(self, ten_file):
        if not ten_file.endswith(".xlsx"):
            ten_file += ".xlsx" 

        current_directory = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(current_directory, ten_file)

        if not os.path.exists(file_path): 
            print(f"Tệp {file_path} không tồn tại.")
            return

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        self.nhan_vien_list = []

        for row in ws.iter_rows(min_row=2, values_only=True): 
            print(f"Đọc dòng: {row}")  
            if len(row) == 3: 
                ma_nv, ten, tuoi = row  
                nhan_vien = NhanVien(ma_nv, ten, tuoi)
                self.them_nhan_vien(nhan_vien)
            else:
                print(f"Dòng dữ liệu không hợp lệ: {row}")  

        print(f"Dữ liệu đã được tải từ: {file_path}")

    def xuat(self):
        print("Danh sách nhân viên:")
        if not self.nhan_vien_list:
            print("Chưa có nhân viên nào.")
            return

        for nv in self.nhan_vien_list:
            print(f"Mã: {nv.ma_nv}, Tên: {nv.ten}, Tuổi: {nv.tuoi}")

if __name__ == "__main__":
    quan_ly = QuanLyNhanVien()

    while True:
        print("\n--- Quản lý nhân viên ---")
        print("1. Thêm nhân viên")
        print("2. Cập nhật nhân viên")
        print("3. Xóa nhân viên")
        print("4. Sắp xếp nhân viên theo tuổi")
        print("5. Lưu vào file Excel")
        print("6. Đọc từ file Excel")
        print("7. Xuất danh sách")
        print("8. Thoát")

        choice = input("Chọn chức năng: ")

        if choice == '1':
            ma_nv = input("Nhập mã nhân viên: ")
            ten_nv = input("Nhập tên nhân viên: ")
            tuoi_nv = int(input("Nhập tuổi nhân viên: "))
            nhan_vien = NhanVien(ma_nv, ten_nv, tuoi_nv)
            quan_ly.them_nhan_vien(nhan_vien)
        elif choice == '2':
            ma_nv = input("Nhập mã nhân viên cần cập nhật: ")
            ten_nv = input("Nhập tên mới (bỏ qua nếu không thay đổi): ")
            tuoi_nv = input("Nhập tuổi mới (bỏ qua nếu không thay đổi): ")
            tuoi_nv = int(tuoi_nv) if tuoi_nv else None
            quan_ly.cap_nhat_nhan_vien(ma_nv, ten_nv, tuoi_nv)
        elif choice == '3':
            ma_nv = input("Nhập mã nhân viên cần xóa: ")
            quan_ly.xoa_nhan_vien(ma_nv)
        elif choice == '4':
            quan_ly.sap_xep_nhan_vien()
            print("Danh sách nhân viên đã được sắp xếp theo tuổi.")
        elif choice == '5':
            ten_file = input("Nhập tên file để lưu: ")
            quan_ly.luu_vao_file(ten_file)
        elif choice == '6':
            ten_file = input("Nhập tên file để đọc: ")
            quan_ly.doc_tu_file(ten_file)
        elif choice == '7':
            quan_ly.xuat()
        elif choice == '8':
            print("Thoát chương trình!")
            break
        else:
            print("Lựa chọn không hợp lệ!")
